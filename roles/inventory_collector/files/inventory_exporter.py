#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Aruba Inventory Exporter

Ce script exporte les données d'inventaire des équipements Aruba vers un fichier Excel.
Il peut être utilisé directement par Ansible ou exécuté manuellement.

Usage:
    python inventory_exporter.py input_file output_file

Arguments:
    input_file:  Chemin vers le fichier JSON contenant les données d'inventaire
    output_file: Chemin vers le fichier Excel de sortie

Auteur: [Votre nom]
Date: [Date de création]
"""

import json
import sys
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import logging
from datetime import datetime

# Configuration du logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class ArubaSwitchInventoryExporter:
    """Classe pour exporter l'inventaire des switches Aruba vers Excel."""
    
    def __init__(self, input_file, output_file):
        """
        Initialiser l'exportateur avec les fichiers source et destination.
        
        Args:
            input_file (str): Chemin vers le fichier JSON d'entrée
            output_file (str): Chemin vers le fichier Excel de sortie
        """
        self.input_file = input_file
        self.output_file = output_file
        self.data = None
        self.column_mapping = {
            'nom_switch': 'Nom du Switch',
            'modele': 'Modèle',
            'serial': 'Numéro de Série',
            'version_os': 'Version OS',
            'date_collecte': 'Date de Collecte',
            'adresse_ip': 'Adresse IP'
        }
        self.column_widths = {
            'Nom du Switch': 25,
            'Modèle': 20,
            'Numéro de Série': 25,
            'Version OS': 20,
            'Date de Collecte': 20,
            'Adresse IP': 15
        }
    
    def load_data(self):
        """Charger les données depuis le fichier JSON."""
        try:
            with open(self.input_file, 'r') as f:
                self.data = json.load(f)
            logger.info(f"Données chargées depuis {self.input_file}")
            logger.info(f"Nombre d'équipements: {len(self.data)}")
            return True
        except (json.JSONDecodeError, FileNotFoundError) as e:
            logger.error(f"Erreur lors du chargement des données: {str(e)}")
            return False
    
    def create_excel(self):
        """Créer le fichier Excel avec formatage avancé."""
        if not self.data:
            logger.error("Aucune donnée à exporter")
            return False
        
        try:
            # Créer DataFrame pandas
            df = pd.DataFrame(self.data)
            
            # Réorganiser et renommer les colonnes
            df = df[[col for col in self.column_mapping.keys() if col in df.columns]]
            df = df.rename(columns=self.column_mapping)
            
            # Créer workbook et feuille
            wb = Workbook()
            ws = wb.active
            ws.title = "Inventaire Aruba"
            
            # Ajouter les en-têtes
            for col_idx, column_name in enumerate(df.columns, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = column_name
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Ajouter les données
            for row_idx, row in enumerate(df.values, start=2):
                for col_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.alignment = Alignment(horizontal='left')
                    
                    # Mettre en évidence les échecs de collecte
                    if value == "ÉCHEC DE COLLECTE":
                        cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            
            # Ajuster la largeur des colonnes
            for col_idx, column_name in enumerate(df.columns, start=1):
                width = self.column_widths.get(column_name, 15)
                ws.column_dimensions[get_column_letter(col_idx)].width = width
            
            # Ajouter une ligne d'information récapitulative
            total_row = len(df) + 3
            ws.cell(row=total_row, column=1).value = "Total des équipements :"
            ws.cell(row=total_row, column=1).font = Font(bold=True)
            ws.cell(row=total_row, column=2).value = len(df)
            
            success_count = sum(1 for row in df.values if "ÉCHEC DE COLLECTE" not in row)
            failure_count = len(df) - success_count
            
            ws.cell(row=total_row+1, column=1).value = "Collectes réussies :"
            ws.cell(row=total_row+1, column=1).font = Font(bold=True)
            ws.cell(row=total_row+1, column=2).value = success_count
            
            ws.cell(row=total_row+2, column=1).value = "Collectes échouées :"
            ws.cell(row=total_row+2, column=1).font = Font(bold=True)
            ws.cell(row=total_row+2, column=2).value = failure_count
            
            # Ajouter les métadonnées d'exportation
            metadata_row = total_row + 4
            ws.cell(row=metadata_row, column=1).value = "Généré le :"
            ws.cell(row=metadata_row, column=2).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Enregistrer le fichier Excel
            wb.save(self.output_file)
            logger.info(f"Fichier exporté avec succès : {self.output_file}")
            
            # Vérifier que le fichier a bien été créé
            if not os.path.exists(self.output_file):
                logger.error(f"Le fichier {self.output_file} n'a pas été créé correctement")
                return False
            
            return True
        
        except Exception as e:
            logger.error(f"Erreur lors de la création du fichier Excel: {str(e)}")
            return False

def main():
    """Point d'entrée principal du script."""
    # Vérifier les arguments
    if len(sys.argv) != 3:
        print(f"Usage: {sys.argv[0]} input_file output_file")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2]
    
    # Exporter les données
    exporter = ArubaSwitchInventoryExporter(input_file, output_file)
    if not exporter.load_data():
        logger.error("Échec du chargement des données")
        sys.exit(1)
    
    if not exporter.create_excel():
        logger.error("Échec de la création du fichier Excel")
        sys.exit(1)
    
    logger.info("Exportation terminée avec succès")
    sys.exit(0)

if __name__ == "__main__":
    main()